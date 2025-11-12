from fastapi import APIRouter, Depends, HTTPException, Body
from sqlalchemy.orm import Session
from typing import List
from sqlalchemy import func, extract
from datetime import datetime, timedelta
from app.db.database import get_db
from app.core.security import get_current_admin_user, get_current_user
from app.models.user import User
from app.models.transaction import Purchase, Sale, Blow, Waste, ExtraExpenditure, SaleLineItem, PurchaseLineItem
from app.models.party import Supplier, Customer
from app.models.item import Stock, Item
from app.models.report import WeeklyReport
from fastapi.responses import StreamingResponse, FileResponse
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
import os
from reportlab.pdfbase import pdfmetrics
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import calendar
import asyncio

router = APIRouter()

@router.get("/balance-sheet")
async def get_balance_sheet(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """Get balance sheet (Admin only)"""
    
    # Accounts Receivable = Total unpaid sales
    accounts_receivable = db.query(func.sum(Sale.total_price)).filter(
        Sale.status != 'cancelled'
    ).scalar()
    accounts_receivable = float(accounts_receivable) if accounts_receivable else 0.0
    
    # Accounts Payable = Total unpaid purchases
    accounts_payable = db.query(func.sum(Purchase.total_amount)).filter(
        Purchase.status != 'cancelled'
    ).scalar()
    accounts_payable = float(accounts_payable) if accounts_payable else 0.0
    
    # Total Sales All Time
    total_sales = db.query(func.sum(Sale.total_price)).filter(
        Sale.status != 'cancelled'
    ).scalar()
    total_sales = float(total_sales) if total_sales else 0.0
    
    # Total Purchases All Time
    total_purchases = db.query(func.sum(Purchase.total_amount)).filter(
        Purchase.status != 'cancelled'
    ).scalar()
    total_purchases = float(total_purchases) if total_purchases else 0.0
    
    # Net Position = Assets - Liabilities
    net_position = total_sales - total_purchases
    
    return {
        "accounts_receivable": accounts_receivable,
        "accounts_payable": accounts_payable,
        "total_sales": total_sales,
        "total_purchases": total_purchases,
        "net_position": net_position
    }


@router.get("/profit")
async def get_profit_report(
    month: int = None,
    year: int = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """Get profit report (Admin only) - FIXED VERSION with line items support"""
    if not month:
        month = datetime.now().month
    if not year:
        year = datetime.now().year
    
    # Calculate sales revenue from SaleLineItem table
    sales_revenue = db.query(func.sum(SaleLineItem.total_price)).join(
        Sale, Sale.bill_number == SaleLineItem.bill_number
    ).filter(
        extract('month', Sale.date) == month,
        extract('year', Sale.date) == year,
    ).scalar()
    sales_revenue = float(sales_revenue) if sales_revenue else 0.0
    
    # Calculate purchase costs from PurchaseLineItem table
    purchase_costs = db.query(func.sum(PurchaseLineItem.total_price)).join(
        Purchase, Purchase.bill_number == PurchaseLineItem.bill_number
    ).filter(
        extract('month', Purchase.date) == month,
        extract('year', Purchase.date) == year,
    ).scalar()
    purchase_costs = float(purchase_costs) if purchase_costs else 0.0
    
    # NOTE: Blow feature is for OPERATIONAL TRACKING of material processing, not profit calculation
    # Blow costs are NOT included in profit calculations
    
    # Calculate waste recovery
    waste_recovery = db.query(func.sum(Waste.total_price)).filter(
        extract('month', Waste.date) == month,
        extract('year', Waste.date) == year
    ).scalar()
    waste_recovery = float(waste_recovery) if waste_recovery else 0.0

    # Calculate extra expenditures
    extra_expenditures = db.query(func.sum(ExtraExpenditure.amount)).filter(
        extract('month', ExtraExpenditure.date) == month,
        extract('year', ExtraExpenditure.date) == year
    ).scalar()
    extra_expenditures = float(extra_expenditures) if extra_expenditures else 0.0

    # Totals and profit
    # Profit = Revenue - (Purchase Costs + Extra Expenditures)
    # NOTE: Blow is operational metric, not a profit cost
    total_revenue = sales_revenue + waste_recovery
    total_costs = purchase_costs + extra_expenditures
    profit = total_revenue - total_costs
    profit_margin = (profit / total_revenue * 100) if total_revenue > 0 else 0.0

    return {
        "month": month,
        "year": year,
        "sales_revenue": sales_revenue,
        "purchase_costs": purchase_costs,
        "waste_recovery": waste_recovery,
        "extra_expenditures": extra_expenditures,
        "total_revenue": total_revenue,
        "total_costs": total_costs,
        "profit": profit,
        "profit_margin": round(profit_margin, 2)
    }
    


@router.post("/pdf/bills")
async def download_multiple_bills(
    bill_numbers: List[str] = Body(..., embed=True),
    bill_type: str = "sale",
    signature_admin: str | None = None,
    signature_ceo: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Generate a combined PDF containing multiple bills with line items."""
    if not bill_numbers:
        raise HTTPException(status_code=400, detail="No bill numbers provided")

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # helper to draw one bill page (without signatures)
    def _draw_bill_page(rec, party_name, bill_no, b_type):
        # margins and header
        margin_left = 36
        margin_right = 36
        content_width = width - margin_left - margin_right
        header_h = 72
        # header band
        band_color = colors.HexColor('#0b69ff')
        c.setFillColor(band_color)
        c.rect(0, height - header_h, width, header_h, stroke=0, fill=1)

        # optional logo (static/logo.png relative to repo root)
        logo_w = 0
        try:
            repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
            logo_path = os.path.join(repo_root, 'static', 'logo.png')
            if os.path.exists(logo_path):
                logo_w = 48
                logo_h = 48
                logo_x = margin_left
                logo_y = height - header_h + (header_h - logo_h) / 2
                c.drawImage(logo_path, logo_x, logo_y, width=logo_w, height=logo_h, mask='auto')
        except Exception:
            logo_w = 0
            pass

        # Left-aligned company branding (name + tagline) to the right of the logo
        left_x = margin_left + (logo_w + 8 if logo_w else 0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 16)
        c.drawString(left_x, height - 36, "Waze Technologies")
        c.setFont("Helvetica", 9)
        c.drawString(left_x, height - 52, "Quality Plastics & Supply")

        # right side: small report metadata (type / bill / date)
        c.setFont("Helvetica-Bold", 10)
        c.drawRightString(width - margin_right - 6, height - 34, f"{b_type.capitalize()}")
        c.setFont("Helvetica", 8)
        c.drawRightString(width - margin_right - 6, height - 48, datetime.now().strftime('%Y-%m-%d'))
        # subtle separator line under header
        c.setStrokeColor(colors.white)
        c.setLineWidth(0.6)
        c.line(margin_left, height - header_h + 2, width - margin_right, height - header_h + 2)
        c.setFont("Helvetica-Bold", 15)
        c.drawRightString(width - margin_right, height - 36, f"{b_type.capitalize()}")
        c.setFont("Helvetica", 10)
        c.drawRightString(width - margin_right, height - 52, f"Bill#: {bill_no}")
        if b_type == 'sale':
            c.drawRightString(width - margin_right, height - 66, f"Sale#: {bill_no}")
        else:
            c.drawRightString(width - margin_right, height - 66, f"Purchase#: {bill_no}")

        # party box
        party_box_y = height - header_h - 20
        party_box_h = 56
        band_color = colors.HexColor('#0b69ff')
        c.setFillColor(band_color)
        c.rect(0, height - band_h, width, band_h, stroke=0, fill=1)

        # optional logo
        logo_w = 0
        try:
            repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
            logo_path = os.path.join(repo_root, 'static', 'logo.png')
            if os.path.exists(logo_path):
                logo_w = 48
                logo_h = 48
                logo_x = margin_left
                logo_y = height - band_h + (band_h - logo_h) / 2
                c.drawImage(logo_path, logo_x, logo_y, width=logo_w, height=logo_h, mask='auto')
        except Exception:
            logo_w = 0
            pass

        # Left-aligned company branding (name + tagline) to the right of the logo
        left_x = margin_left + (logo_w + 8 if logo_w else 0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 16)
        c.drawString(left_x, height - 36, "Waze Technologies")
        c.setFont("Helvetica", 9)
        c.drawString(left_x, height - 52, "Quality Plastics & Supply")

        # Report title and page (right-aligned)
        c.setFont("Helvetica-Bold", 11)
        c.drawRightString(width - margin_right - 6, height - 34, f"Combined {bill_type.capitalize()} Report")
        c.setFont("Helvetica", 8)
        c.drawRightString(width - margin_right - 6, height - 48, f"Page {page_num}    {datetime.now().strftime('%Y-%m-%d')}")

        # subtle separator line under header
        c.setStrokeColor(colors.white)
        c.setLineWidth(0.6)
        c.line(margin_left, height - band_h + 2, width - margin_right, height - band_h + 2)
        c.rect(table_x, table_y, table_w, row_h, stroke=0, fill=1)
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(table_x + 6, table_y + 6, "Item")
        c.drawString(table_x + 180, table_y + 6, "Quantity")
        c.drawString(table_x + 300, table_y + 6, "Unit Price")
        c.drawString(table_x + 420, table_y + 6, "Total")

        # item row
        item_row_y = table_y - row_h
        c.setFillColor(colors.white)
        c.rect(table_x, item_row_y, table_w, row_h, stroke=0, fill=1)
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 11)
        c.drawString(table_x + 6, item_row_y + 6, str(rec.item_id))
        c.drawString(table_x + 180, item_row_y + 6, str(rec.quantity))
        unit_price_val = float(getattr(rec, 'unit_price', 0) or 0)
        c.drawRightString(table_x + 400, item_row_y + 6, f"{unit_price_val:,.2f}")
        total_field = float(getattr(rec, 'total_amount', None) or getattr(rec, 'total_price', 0) or 0)
        c.drawRightString(table_x + table_w - 8, item_row_y + 6, f"{total_field:,.2f}")

        # totals
        totals_y_local = item_row_y - 44
        c.setFont("Helvetica-Bold", 12)
        c.drawRightString(width - margin_right - 60, totals_y_local + 14, "Grand Total:")
        c.setFillColor(colors.HexColor('#0b69ff'))
        c.setFont("Helvetica-Bold", 12)
        c.drawRightString(width - margin_right, totals_y_local + 14, f"{total_field:,.2f}")

        # watermark
        c.saveState()
        c.setFillColor(colors.HexColor('#e9f3ff'))
        c.setFont("Helvetica-Bold", 60)
        c.translate(width/2, height/2)
        c.rotate(30)
        c.drawCentredString(0, 0, "WAZE ENTERPRISES")
        c.restoreState()

        return

    # Collect valid records with line items
    records = []
    for bill_no in bill_numbers:
        if bill_type == 'purchase':
            bill = db.query(Purchase).filter(Purchase.bill_number == bill_no).first()
            if not bill:
                continue
            party = db.query(Supplier).filter(Supplier.id == bill.supplier_id).first()
            party_name = party.name if party else str(bill.supplier_id)
            # Get line items for this purchase
            line_items = db.query(PurchaseLineItem).filter(
                PurchaseLineItem.bill_number == bill_no
            ).all()
        else:
            bill = db.query(Sale).filter(Sale.bill_number == bill_no).first()
            if not bill:
                continue
            party = db.query(Customer).filter(Customer.id == bill.customer_id).first()
            party_name = party.name if party else str(bill.customer_id)
            # Get line items for this sale
            line_items = db.query(SaleLineItem).filter(
                SaleLineItem.bill_number == bill_no
            ).all()

        if line_items:
            for line_item in line_items:
                # Get item name
                item = db.query(Item).filter(Item.id == line_item.item_id).first()
                item_name = item.name if item else str(line_item.item_id)
                
                records.append({
                    'bill_no': bill_no,
                    'party_name': party_name,
                    'item_name': item_name,
                    'quantity': line_item.quantity,
                    'unit_price': float(line_item.unit_price) if line_item.unit_price else 0,
                    'total_price': float(line_item.total_price) if line_item.total_price else 0
                })

    if not records:
        raise HTTPException(status_code=404, detail="No matching bills found")

    # Pre-calculate grand total
    grand_total = sum(r['total_price'] for r in records)

    # Table layout parameters
    margin_left = 36
    margin_right = 36
    top_margin = 120
    bottom_margin = 72
    usable_height = height - top_margin - bottom_margin

    # table row heights
    header_h = 20
    # increase row height to reduce vertical crowding and avoid text overlapping grid lines
    row_h = 28
    # compute number of rows per page (reserve space for header and some footer)
    rows_per_page = max(3, int((usable_height - header_h) // row_h))

    # compute precise column widths to fit content_w
    content_x = margin_left
    content_w = width - margin_left - margin_right
    # Remove Date column (sales bill_number already encodes date)
    # Reduce Bill# width to free space for Total (user requested maximum space for Total)
    # New approx base widths: Bill# 120, Customer 140, Item 170, Qty 40, Unit Price 80, Total 190
    # Transfer 10% of Total column width to Bill# column (10% of 190 ~= 19)
    base_col_w = [139, 140, 170, 40, 80, 171]  # Bill#, Customer, Item, Qty, Unit Price, Total
    scale = content_w / sum(base_col_w)
    col_widths = [w * scale for w in base_col_w]
    # x positions (left edges) for columns
    col_x = [content_x]
    for w in col_widths:
        col_x.append(col_x[-1] + w)
    # ensure final right edge exact
    col_x[-1] = content_x + content_w

    def _draw_page_header(page_num):
        # top header (company band)
        band_h = 72
        c.setFillColor(colors.HexColor('#0b69ff'))
        c.rect(0, height - band_h, width, band_h, stroke=0, fill=1)
        # Centered company branding in the header
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 18)
        c.drawString(margin_left, height - 25, "Waze Technologies")
        c.setFont("Helvetica", 10)
        c.drawString(margin_left, height - 41, "Rawat Industrial Area")
        c.drawString(margin_left, height - 55, "03439998954")

        # Report title and page
        c.setFont("Helvetica-Bold", 12)
        #c.drawRightString(width - margin_right, height - 36, f"Combined {bill_type.capitalize()} Report")
        c.setFont("Helvetica", 9)
        #c.drawRightString(width - margin_right, height - 52, f"Page {page_num}")

        # draw table header
        y_header = height - top_margin + 8
        c.setFillColor(colors.HexColor('#cfe3ff'))
        c.rect(content_x, y_header - header_h + 4, content_w, header_h, stroke=0, fill=1)
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 10)
        # column titles (removed Date column) — rename Party to Customer and shorten Unit Price
        titles = ["Bill#", "Customer", "Item", "Qty", "Unit", "Total"]
        for i, title in enumerate(titles):
            x0 = col_x[i]
            x1 = col_x[i+1] if i+1 < len(col_x) else content_x + content_w
            # fit header text if necessary
            max_w = x1 - x0 - 8
            txt = title
            while pdfmetrics.stringWidth(txt, "Helvetica-Bold", 10) > max_w and len(txt) > 1:
                txt = txt[:-1]
            if pdfmetrics.stringWidth(title, "Helvetica-Bold", 10) > max_w:
                txt = txt[:-1] + "…"
            c.drawString(x0 + 4, y_header - 14, txt)

        return y_header - header_h

    # draw rows with pagination
    total = 0
    page = 1
    idx = 0
    while idx < len(records):
        # Draw header for this page
        row_start_y = _draw_page_header(page)

        # draw rows (with grid lines and fitted text)
        rows_drawn = 0
        for r in range(rows_per_page):
            if idx >= len(records):
                break
            rec = records[idx]
            # top y for this row block
            y_top = row_start_y - (r * row_h)
            # place text baseline a bit above the vertical center of the row
            y_text = y_top - (row_h / 2) + 4
            c.setFont("Helvetica", 9)
            c.setFillColor(colors.black)

            # draw horizontal separator line for this row (a few points below the row)
            y_line = y_top - row_h
            c.setStrokeColor(colors.HexColor('#d9e6fb'))
            c.setLineWidth(0.5)
            c.line(margin_left, y_line + 2, width - margin_right, y_line + 2)

            # draw vertical grid lines (use col_x positions) from header down to this row's bottom
            c.setStrokeColor(colors.HexColor('#bfcfe8'))
            grid_top = row_start_y + 8
            grid_bottom = y_line + 2
            for xg in col_x:
                c.line(xg, grid_top, xg, grid_bottom)
            # rightmost vertical line
            c.line(width - margin_right, grid_top, width - margin_right, grid_bottom)

            # helper to fit text into a cell width
            def fit_and_draw(text, x0, x1, y_pos, align='left'):
                max_w = x1 - x0 - 8
                text = '' if text is None else str(text)
                if pdfmetrics.stringWidth(text, 'Helvetica', 9) <= max_w:
                    if align == 'left':
                        c.drawString(x0 + 4, y_pos, text)
                    else:
                        c.drawRightString(x1 - 6, y_pos, text)
                    return
                # truncate
                t = text
                while t and pdfmetrics.stringWidth(t + '…', 'Helvetica', 9) > max_w:
                    t = t[:-1]
                if align == 'left':
                    c.drawString(x0 + 4, y_pos, t + '…')
                else:
                    c.drawRightString(x1 - 6, y_pos, t + '…')

            # prepare cell boundaries
            cell_bounds = []
            for i, x0 in enumerate(col_x):
                if i + 1 < len(col_x):
                    x1 = col_x[i + 1]
                else:
                    x1 = width - margin_right
                cell_bounds.append((x0, x1))

            # Draw data from record dict
            fit_and_draw(rec['bill_no'], cell_bounds[0][0], cell_bounds[0][1], y_text, 'left')
            fit_and_draw(rec['party_name'], cell_bounds[1][0], cell_bounds[1][1], y_text, 'left')
            fit_and_draw(rec['item_name'], cell_bounds[2][0], cell_bounds[2][1], y_text, 'left')
            fit_and_draw(str(rec['quantity']), cell_bounds[3][0], cell_bounds[3][1], y_text, 'right')
            fit_and_draw(f"{rec['unit_price']:,.2f}", cell_bounds[4][0], cell_bounds[4][1], y_text, 'right')
            fit_and_draw(f"{rec['total_price']:,.2f}", cell_bounds[5][0], cell_bounds[5][1], y_text, 'right')

            total += rec['total_price']
            idx += 1
            rows_drawn += 1

        # draw page subtotal immediately after the table end (so subtotal sits right below the rows)
        y_table_end = row_start_y - (rows_drawn * row_h) - 8
        c.setFont("Helvetica-Bold", 10)
        #c.drawRightString(width - margin_right, y_table_end - 6, f"Page subtotal: {total:,.2f}")

        # If more records remain, create new page and continue
        if idx < len(records):
            c.showPage()
            page += 1
            continue
        else:
            # final page: draw grand total right after the table end
            c.setFont("Helvetica-Bold", 12)
            c.setFillColor(colors.black)
            grand_total_y = y_table_end - 26
            c.drawRightString(width - margin_right, grand_total_y, f"Grand Total: {grand_total:,.2f}")

            # Signature area below grand total (on the same page)
            sig_w = 120
            sig_h = 48
            # add extra gap between grand total and signatures to avoid overlap
            extra_gap = 100
            sig_y = grand_total_y - extra_gap - sig_h
            c.setFont("Helvetica", 15)  # Increased from 9 to 15 (added 6px)

            # Admin signature (left)
            admin_x = margin_left
            c.drawString(admin_x, sig_y + sig_h + 8, "Authorized by Waheed")
            if signature_admin and os.path.exists(signature_admin):
                try:
                    c.drawImage(signature_admin, admin_x, sig_y, width=sig_w, height=sig_h, mask='auto')
                except Exception:
                    pass

            # CEO signature (right)
            ceo_x = width - margin_right - sig_w
            c.drawString(ceo_x, sig_y + sig_h + 8, "Authorized by Zeeshan")
            if signature_ceo and os.path.exists(signature_ceo):
                try:
                    c.drawImage(signature_ceo, ceo_x, sig_y, width=sig_w, height=sig_h, mask='auto')
                except Exception:
                    pass

            break

    # Draw final grand total and signatures on the last page (below the table)
    # Grand total (right aligned above signature area) — use pre-calculated grand_total across all bills
    # grand_total_y = bottom_margin + 36
    # c.setFont("Helvetica-Bold", 12)
    # c.setFillColor(colors.black)
    # c.drawRightString(width - margin_right, grand_total_y, f"Grand Total: {grand_total:,.2f}")

    # # Signature area below grand total
    # sig_w = 120
    # sig_h = 48
    # sig_y = bottom_margin - sig_h - 8
    # c.setFont("Helvetica", 9)

    # # Admin signature (left)
    # admin_x = margin_left
    # c.drawString(admin_x, sig_y + sig_h + 8, "Admin Signature")
    # if signature_admin and os.path.exists(signature_admin):
    #     try:
    #         c.drawImage(signature_admin, admin_x, sig_y, width=sig_w, height=sig_h, mask='auto')
    #     except Exception:
    #         pass

    # # CEO signature (right)
    # ceo_x = width - margin_right - sig_w
    # c.drawString(ceo_x, sig_y + sig_h + 8, "CEO Signature")
    # if signature_ceo and os.path.exists(signature_ceo):
    #     try:
    #         c.drawImage(signature_ceo, ceo_x, sig_y, width=sig_w, height=sig_h, mask='auto')
    #     except Exception:
    #         pass

    c.save()
    buffer.seek(0)
    filename = f"combined-{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=\"{filename}\""})


@router.get("/export-excel")
async def export_weekly_report_excel(
    week_offset: int = 0,
    date: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """
    Export report as Excel file.
    - If date parameter provided (YYYY-MM-DD): exports data for that specific day only
    - Otherwise: exports weekly report using week_offset (0=current week, -1=last week, etc)
    """
    
    if date:
        # Daily report for specific date
        try:
            from datetime import date as date_class
            report_date = datetime.strptime(date, '%Y-%m-%d').date()
            day_start = datetime.combine(report_date, datetime.min.time())
            day_end = day_start + timedelta(days=1)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    else:
        # Weekly report
        today = datetime.now()
        # ISO week: Monday=0, Sunday=6
        days_since_monday = today.weekday()
        day_start = today - timedelta(days=days_since_monday) - timedelta(weeks=abs(week_offset))
        day_start = day_start.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=7)
        report_date = None

    # Calculate sales revenue
    sales_revenue = db.query(func.sum(Sale.total_price)).filter(
        Sale.date >= day_start,
        Sale.date < day_end,
        Sale.status != 'cancelled'
    ).scalar()
    sales_revenue = float(sales_revenue) if sales_revenue else 0.0

    # Calculate purchase costs
    purchase_costs = db.query(func.sum(Purchase.total_amount)).filter(
        Purchase.date >= day_start,
        Purchase.date < day_end,
        Purchase.status != 'cancelled'
    ).scalar()
    purchase_costs = float(purchase_costs) if purchase_costs else 0.0

    # NOTE: Blow feature is for OPERATIONAL TRACKING, not profit calculation
    # Blow costs are NOT included in profit calculations
    
    # Calculate waste recovery
    waste_recovery = db.query(func.sum(Waste.total_price)).filter(
        Waste.date >= day_start,
        Waste.date < day_end
    ).scalar()
    waste_recovery = float(waste_recovery) if waste_recovery else 0.0

    # Calculate extra expenditures
    extra_expenditures = db.query(func.sum(ExtraExpenditure.amount)).filter(
        ExtraExpenditure.date >= day_start,
        ExtraExpenditure.date < day_end
    ).scalar()
    extra_expenditures = float(extra_expenditures) if extra_expenditures else 0.0

    # Totals
    total_revenue = sales_revenue + waste_recovery
    total_costs = purchase_costs + extra_expenditures
    profit = total_revenue - total_costs
    profit_margin = (profit / total_revenue * 100) if total_revenue > 0 else 0.0

    # Get detailed transactions for daily report
    if date:
        # Fetch detailed sales
        sales = db.query(Sale).filter(
            Sale.date >= day_start,
            Sale.date < day_end,
            Sale.status != 'cancelled'
        ).all()
        
        # Fetch detailed purchases
        purchases = db.query(Purchase).filter(
            Purchase.date >= day_start,
            Purchase.date < day_end,
            Purchase.status != 'cancelled'
        ).all()
        
        # Fetch detailed waste
        waste_items = db.query(Waste).filter(
            Waste.date >= day_start,
            Waste.date < day_end
        ).all()
        
        # Fetch detailed extra expenditures
        extra_expenditure_items = db.query(ExtraExpenditure).filter(
            ExtraExpenditure.date >= day_start,
            ExtraExpenditure.date < day_end
        ).all()
    else:
        sales = []
        purchases = []
        waste_items = []
        extra_expenditure_items = []

    # Count transactions
    sales_count = db.query(func.count(Sale.bill_number)).filter(
        Sale.date >= day_start,
        Sale.date < day_end,
        Sale.status != 'cancelled'
    ).scalar() or 0
    
    purchase_count = db.query(func.count(Purchase.bill_number)).filter(
        Purchase.date >= day_start,
        Purchase.date < day_end,
        Purchase.status != 'cancelled'
    ).scalar() or 0

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Define styles
    header_fill = PatternFill(start_color="0073E6", end_color="0073E6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14, color="0073E6")
    section_fill = PatternFill(start_color="E9F3FF", end_color="E9F3FF", fill_type="solid")
    section_font = Font(bold=True, color="000000", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws['A1'] = "WAZE TECHNOLOGIES - DAILY PROFIT REPORT" if date else "WAZE TECHNOLOGIES - WEEKLY PROFIT REPORT"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    # Period
    if date:
        period_text = f"Date: {report_date.strftime('%B %d, %Y')}"
    else:
        week_period = f"Week of {day_start.strftime('%B %d, %Y')} to {(day_end - timedelta(days=1)).strftime('%B %d, %Y')}"
        period_text = week_period
    ws['A2'] = period_text
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:F2')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 18

    # Revenue section (row 4)
    ws['A4'] = "REVENUE"
    ws['A4'].fill = section_fill
    ws['A4'].font = section_font
    ws.merge_cells('A4:F4')
    ws.row_dimensions[4].height = 20

    row = 5
    revenue_items = [
        ("Sales Revenue", sales_revenue),
        ("Waste Recovery", waste_recovery),
        ("Total Revenue", total_revenue)
    ]
    for label, value in revenue_items:
        ws[f'A{row}'] = label
        ws[f'F{row}'] = value
        ws[f'F{row}'].number_format = '#,##0.00'
        if label == "Total Revenue":
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'F{row}'].font = Font(bold=True)
        row += 1

    # Costs section
    row += 1
    ws[f'A{row}'] = "COSTS"
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws.merge_cells(f'A{row}:F{row}')
    ws.row_dimensions[row].height = 20
    row += 1

    cost_items = [
        ("Purchase Costs", purchase_costs),
        ("Extra Expenditures", extra_expenditures),
        ("Total Costs", total_costs)
    ]
    for label, value in cost_items:
        ws[f'A{row}'] = label
        ws[f'F{row}'] = value
        ws[f'F{row}'].number_format = '#,##0.00'
        if label == "Total Costs":
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'F{row}'].font = Font(bold=True)
        row += 1

    # Profit section
    row += 1
    ws[f'A{row}'] = "PROFIT"
    ws[f'A{row}'].fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    ws[f'A{row}'].font = Font(bold=True, size=11, color="2E7D32")
    ws.merge_cells(f'A{row}:F{row}')
    ws.row_dimensions[row].height = 20
    row += 1

    profit_items = [
        ("Gross Profit", profit),
        ("Profit Margin (%)", profit_margin)
    ]
    for label, value in profit_items:
        ws[f'A{row}'] = label
        ws[f'F{row}'] = value
        if '%' in label:
            ws[f'F{row}'].number_format = '0.00"%"'
        else:
            ws[f'F{row}'].number_format = '#,##0.00'
        ws[f'A{row}'].font = Font(bold=True, color="2E7D32")
        ws[f'F{row}'].font = Font(bold=True, color="2E7D32")
        row += 1

    # Summary section
    row += 1
    ws[f'A{row}'] = "TRANSACTION SUMMARY"
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws.merge_cells(f'A{row}:F{row}')
    ws.row_dimensions[row].height = 20
    row += 1

    summary_items = [
        ("Total Sales Transactions", sales_count),
        ("Total Purchase Transactions", purchase_count)
    ]
    for label, value in summary_items:
        ws[f'A{row}'] = label
        ws[f'F{row}'] = value
        ws[f'F{row}'].number_format = '0'
        row += 1

    # Detailed transactions section (for daily reports only)
    if date and (sales or purchases or waste_items or extra_expenditure_items):
        row += 2
        ws[f'A{row}'] = "DETAILED TRANSACTIONS"
        ws[f'A{row}'].fill = section_fill
        ws[f'A{row}'].font = section_font
        ws.merge_cells(f'A{row}:F{row}')
        ws.row_dimensions[row].height = 20
        row += 1

        # Sales transactions
        if sales:
            ws[f'A{row}'] = "SALES"
            ws[f'A{row}'].font = Font(bold=True, color="2E7D32")
            row += 1
            
            # Headers
            headers = ["Bill Number", "Customer", "Items", "Total", "Status", "Date"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = PatternFill(start_color="B3E5FC", end_color="B3E5FC", fill_type="solid")
                cell.font = Font(bold=True)
            row += 1
            
            for sale in sales:
                ws[f'A{row}'] = sale.bill_number
                # Get customer name - handle relationship not being loaded
                if hasattr(sale, 'customer') and sale.customer:
                    customer_name = sale.customer.name
                else:
                    # Fetch customer from DB
                    customer = db.query(Customer).filter(Customer.id == sale.customer_id).first()
                    customer_name = customer.name if customer else f"Customer {sale.customer_id}"
                ws[f'B{row}'] = customer_name
                
                # Get items for this sale
                items_str = ", ".join([f"{li.item.name} x {li.quantity}" for li in sale.line_items]) if sale.line_items else "N/A"
                ws[f'C{row}'] = items_str
                ws[f'D{row}'] = sale.total_price
                ws[f'D{row}'].number_format = '#,##0.00'
                ws[f'E{row}'] = sale.payment_status or "pending"
                ws[f'F{row}'] = sale.date.strftime('%Y-%m-%d') if sale.date else ""
                row += 1

        # Purchase transactions
        if purchases:
            row += 1
            ws[f'A{row}'] = "PURCHASES"
            ws[f'A{row}'].font = Font(bold=True, color="C62828")
            row += 1
            
            # Headers
            headers = ["Bill Number", "Supplier", "Items", "Total", "Status", "Date"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = PatternFill(start_color="FFCCBC", end_color="FFCCBC", fill_type="solid")
                cell.font = Font(bold=True)
            row += 1
            
            for purchase in purchases:
                ws[f'A{row}'] = purchase.bill_number
                # Get supplier name - handle relationship not being loaded
                if hasattr(purchase, 'supplier') and purchase.supplier:
                    supplier_name = purchase.supplier.name
                else:
                    # Fetch supplier from DB
                    supplier = db.query(Supplier).filter(Supplier.id == purchase.supplier_id).first()
                    supplier_name = supplier.name if supplier else f"Supplier {purchase.supplier_id}"
                ws[f'B{row}'] = supplier_name
                
                # Get items for this purchase
                items_str = ", ".join([f"{li.item.name} x {li.quantity}" for li in purchase.line_items]) if purchase.line_items else "N/A"
                ws[f'C{row}'] = items_str
                ws[f'D{row}'] = purchase.total_amount
                ws[f'D{row}'].number_format = '#,##0.00'
                ws[f'E{row}'] = purchase.payment_status or "pending"
                ws[f'F{row}'] = purchase.date.strftime('%Y-%m-%d') if purchase.date else ""
                row += 1

        # Waste transactions
        if waste_items:
            row += 1
            ws[f'A{row}'] = "WASTE RECOVERY"
            ws[f'A{row}'].font = Font(bold=True, color="F57F17")
            row += 1
            
            # Headers
            headers = ["ID", "Item", "Quantity", "Price/Unit", "Total", "Date"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                cell.font = Font(bold=True)
            row += 1
            
            for waste in waste_items:
                ws[f'A{row}'] = waste.id
                # Get item name - handle relationship not being loaded
                if hasattr(waste, 'item') and waste.item:
                    item_name = waste.item.name
                else:
                    # Fetch item from DB
                    item = db.query(Item).filter(Item.id == waste.item_id).first()
                    item_name = item.name if item else f"Item {waste.item_id}"
                ws[f'B{row}'] = item_name
                ws[f'C{row}'] = waste.quantity
                ws[f'D{row}'] = waste.price_per_unit
                ws[f'D{row}'].number_format = '#,##0.00'
                ws[f'E{row}'] = waste.total_price
                ws[f'E{row}'].number_format = '#,##0.00'
                ws[f'F{row}'] = waste.date.strftime('%Y-%m-%d') if waste.date else ""
                row += 1

        # Extra expenditures transactions
        if extra_expenditure_items:
            row += 1
            ws[f'A{row}'] = "EXTRA EXPENDITURES"
            ws[f'A{row}'].font = Font(bold=True, color="9C27B0")
            row += 1
            
            # Headers
            headers = ["ID", "Description", "Category", "Amount", "Notes", "Date"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid")
                cell.font = Font(bold=True)
            row += 1
            
            for exp in extra_expenditure_items:
                ws[f'A{row}'] = exp.id
                ws[f'B{row}'] = exp.description or ""
                ws[f'C{row}'] = exp.category or ""
                ws[f'D{row}'] = exp.amount
                ws[f'D{row}'].number_format = '#,##0.00'
                ws[f'E{row}'] = exp.notes or ""
                ws[f'F{row}'] = exp.date.strftime('%Y-%m-%d') if exp.date else ""
                row += 1

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['F'].width = 18

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Generate filename
    if date:
        filename = f"daily-report-{date}.xlsx"
    else:
        filename = f"weekly-report-{day_start.strftime('%Y-%m-%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )


@router.post("/generate-weekly")
async def generate_weekly_report(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """
    Generate and save a weekly report to the database.
    Called by scheduler or manually.
    """
    # Get last week's data
    today = datetime.now()
    days_since_monday = today.weekday()
    week_start = today - timedelta(days=days_since_monday) - timedelta(weeks=1)
    week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)
    week_end = week_start + timedelta(days=7)

    # Calculate sales revenue
    sales_revenue = db.query(func.sum(Sale.total_price)).filter(
        Sale.date >= week_start,
        Sale.date < week_end,
        Sale.status != 'cancelled'
    ).scalar()
    sales_revenue = float(sales_revenue) if sales_revenue else 0.0

    # Calculate purchase costs
    purchase_costs = db.query(func.sum(Purchase.total_amount)).filter(
        Purchase.date >= week_start,
        Purchase.date < week_end,
        Purchase.status != 'cancelled'
    ).scalar()
    purchase_costs = float(purchase_costs) if purchase_costs else 0.0

    # Calculate blow costs
    blow_costs = db.query(func.sum(Blow.blow_cost_per_unit * Blow.quantity)).filter(
        Blow.date_time >= week_start,
        Blow.date_time < week_end
    ).scalar()
    blow_costs = float(blow_costs) if blow_costs else 0.0

    # Calculate sales blow price costs
    sales_blow_costs = db.query(func.sum(Sale.blow_price * Sale.quantity)).filter(
        Sale.date >= week_start,
        Sale.date < week_end,
        Sale.status != 'cancelled'
    ).scalar()
    sales_blow_costs = float(sales_blow_costs) if sales_blow_costs else 0.0

    # Calculate waste recovery
    waste_recovery = db.query(func.sum(Waste.total_price)).filter(
        Waste.date >= week_start,
        Waste.date < week_end
    ).scalar()
    waste_recovery = float(waste_recovery) if waste_recovery else 0.0

    # Calculate extra expenditures
    extra_expenditures = db.query(func.sum(ExtraExpenditure.amount)).filter(
        ExtraExpenditure.date >= week_start,
        ExtraExpenditure.date < week_end
    ).scalar()
    extra_expenditures = float(extra_expenditures) if extra_expenditures else 0.0

    # Totals
    total_revenue = sales_revenue + waste_recovery
    total_costs = purchase_costs + blow_costs + sales_blow_costs + extra_expenditures
    profit = total_revenue - total_costs
    profit_margin = (profit / total_revenue * 100) if total_revenue > 0 else 0.0

    # Count transactions
    total_transactions = db.query(func.count(Sale.bill_number)).filter(
        Sale.date >= week_start,
        Sale.date < week_end,
        Sale.status != 'cancelled'
    ).scalar() or 0
    total_transactions += db.query(func.count(Purchase.bill_number)).filter(
        Purchase.date >= week_start,
        Purchase.date < week_end,
        Purchase.status != 'cancelled'
    ).scalar() or 0

    # Create summary
    summary = f"Sales: PKR {sales_revenue:,.0f} | Costs: PKR {total_costs:,.0f} | Profit: PKR {profit:,.0f}"

    # Get ISO week info
    iso_calendar = week_start.isocalendar()
    week_number = iso_calendar.week
    year = iso_calendar.year

    # Check if report already exists
    existing = db.query(WeeklyReport).filter(
        WeeklyReport.year == year,
        WeeklyReport.week_number == week_number
    ).first()

    if existing:
        # Update existing
        existing.sales_revenue = sales_revenue
        existing.purchase_costs = purchase_costs
        existing.blow_costs = blow_costs
        existing.waste_recovery = waste_recovery
        existing.total_revenue = total_revenue
        existing.total_costs = total_costs
        existing.profit = profit
        existing.profit_margin = profit_margin
        existing.total_transactions = total_transactions
        existing.summary = summary
    else:
        # Create new
        report = WeeklyReport(
            week_start=week_start,
            week_end=week_end,
            year=year,
            week_number=week_number,
            sales_revenue=sales_revenue,
            purchase_costs=purchase_costs,
            blow_costs=blow_costs,
            waste_recovery=waste_recovery,
            total_revenue=total_revenue,
            total_costs=total_costs,
            profit=profit,
            profit_margin=profit_margin,
            total_transactions=total_transactions,
            summary=summary,
            generated_by=current_user.id
        )
        db.add(report)

    db.commit()

    return {
        "message": "Weekly report generated successfully",
        "week": f"{week_start.strftime('%Y-%m-%d')} to {week_end.strftime('%Y-%m-%d')}",
        "profit": profit,
        "revenue": total_revenue
    }


@router.get("/weekly-reports")
async def get_weekly_reports(
    limit: int = 12,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """Get list of stored weekly reports (last 12 weeks by default)"""
    try:
        reports = db.query(WeeklyReport).order_by(
            WeeklyReport.year.desc(),
            WeeklyReport.week_number.desc()
        ).limit(limit).all()

        return {
            "total": len(reports),
            "reports": [
                {
                    "id": r.id,
                    "week_start": r.week_start.isoformat(),
                    "week_end": r.week_end.isoformat(),
                    "year": r.year,
                    "week_number": r.week_number,
                    "profit": r.profit,
                    "revenue": r.total_revenue,
                    "margin": r.profit_margin,
                    "transactions": r.total_transactions,
                    "summary": r.summary,
                    "created_at": r.created_at.isoformat() if r.created_at else None
                }
                for r in reports
            ]
        }
    except Exception as e:
        # Return empty list if table doesn't exist or other errors
        return {
            "total": 0,
            "reports": [],
            "message": "No weekly reports generated yet"
        }


@router.get("/export-purchases-excel")
async def export_purchases_excel(
    bill_numbers: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """
    Export purchases as Excel file.
    If bill_numbers provided (comma-separated), export only those.
    Otherwise export all purchases.
    """
    try:
        # Get purchases
        if bill_numbers:
            bill_list = [b.strip() for b in bill_numbers.split(',')]
            purchases = db.query(Purchase).filter(Purchase.bill_number.in_(bill_list)).all()
        else:
            purchases = db.query(Purchase).all()

        # Get suppliers and items for lookups
        suppliers_map = {s.id: s.name for s in db.query(Supplier).all()}
        items_map = {i.id: i.name for i in db.query(Item).all()}

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Purchases"

        # Define styles
        header_fill = PatternFill(start_color="0073E6", end_color="0073E6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        title_font = Font(bold=True, size=14, color="0073E6")
        total_fill = PatternFill(start_color="E9F3FF", end_color="E9F3FF", fill_type="solid")

        # Title
        ws['A1'] = "PURCHASE ORDERS EXPORT"
        ws['A1'].font = title_font
        ws.merge_cells('A1:I1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25

        # Export date
        export_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A2'] = f"Exported on: {export_date}"
        ws['A2'].font = Font(size=9, italic=True)
        ws.merge_cells('A2:I2')
        ws.row_dimensions[2].height = 18

        # Headers (Row 4)
        headers = [
            "Bill Number",
            "Supplier",
            "Item",
            "Quantity",
            "Unit Price",
            "Total Amount",
            "Payment Status",
            "Paid Amount",
            "Date"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[4].height = 20

        # Data rows
        row = 5
        total_amount_sum = 0
        paid_amount_sum = 0

        for purchase in purchases:
            supplier_name = suppliers_map.get(purchase.supplier_id, purchase.supplier_id)
            
            # Get line items for this purchase
            line_items = db.query(PurchaseLineItem).filter(PurchaseLineItem.bill_number == purchase.bill_number).all()
            
            if line_items:
                for line_item in line_items:
                    item_name = items_map.get(line_item.item_id, line_item.item_id)
                    
                    cells_data = [
                        purchase.bill_number,
                        supplier_name,
                        item_name,
                        line_item.quantity,
                        float(line_item.unit_price) if line_item.unit_price else 0,
                        float(line_item.total_price) if line_item.total_price else 0,
                        purchase.payment_status,
                        float(purchase.paid_amount) if purchase.paid_amount else 0,
                        purchase.date.strftime("%Y-%m-%d") if purchase.date else ""
                    ]

                    for col, value in enumerate(cells_data, 1):
                        cell = ws.cell(row=row, column=col)
                        cell.value = value
                        cell.border = border
                        cell.alignment = Alignment(horizontal='left', vertical='center')

                        # Format currency columns
                        if col in [5, 6, 8]:  # Unit Price, Total Amount, Paid Amount
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal='right', vertical='center')

                    # Sum totals
                    total_amount_sum += float(line_item.total_price) if line_item.total_price else 0
                    paid_amount_sum += float(purchase.paid_amount) if purchase.paid_amount else 0
                    row += 1

        # Summary row
        row += 1
        ws[f'A{row}'] = "SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'A{row}'].fill = total_fill
        ws.merge_cells(f'A{row}:I{row}')
        ws.row_dimensions[row].height = 20

        row += 1
        ws[f'A{row}'] = "Total Records:"
        ws[f'B{row}'] = len(purchases)
        ws[f'A{row}'].fill = total_fill
        ws[f'B{row}'].fill = total_fill

        row += 1
        ws[f'A{row}'] = "Total Purchase Amount:"
        ws[f'B{row}'] = total_amount_sum
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'A{row}'].fill = total_fill
        ws[f'B{row}'].fill = total_fill

        row += 1
        ws[f'A{row}'] = "Total Paid:"
        ws[f'B{row}'] = paid_amount_sum
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'A{row}'].fill = total_fill
        ws[f'B{row}'].fill = total_fill

        row += 1
        ws[f'A{row}'] = "Outstanding Balance:"
        ws[f'B{row}'] = total_amount_sum - paid_amount_sum
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'A{row}'].fill = total_fill
        ws[f'B{row}'].fill = total_fill

        # Column widths
        column_widths = [28, 20, 20, 12, 12, 15, 12, 15, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Generate filename
        if bill_numbers:
            filename = f"purchases-selected-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        else:
            filename = f"purchases-all-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export-sales-excel")
async def export_sales_excel(
    bill_numbers: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """
    Export sales as Excel file.
    If bill_numbers provided (comma-separated), export only those.
    Otherwise export all sales.
    """
    try:
        # Get sales
        if bill_numbers:
            bill_list = [b.strip() for b in bill_numbers.split(',')]
            sales = db.query(Sale).filter(Sale.bill_number.in_(bill_list)).all()
        else:
            sales = db.query(Sale).all()

        # Get customers and items for lookups
        customers_map = {c.id: c.name for c in db.query(Customer).all()}
        items_map = {i.id: i.name for i in db.query(Item).all()}

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales"

        # Define styles
        header_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        title_font = Font(bold=True, size=14, color="00B050")
        total_fill = PatternFill(start_color="E2EFD9", end_color="E2EFD9", fill_type="solid")

        # Title
        ws['A1'] = "SALES EXPORT"
        ws['A1'].font = title_font
        ws.merge_cells('A1:J1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25

        # Export date
        export_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A2'] = f"Exported on: {export_date}"
        ws['A2'].font = Font(size=9, italic=True)
        ws.merge_cells('A2:J2')
        ws.row_dimensions[2].height = 18

        # Headers (Row 4)
        headers = [
            "Bill Number",
            "Customer",
            "Item",
            "Quantity",
            "Unit Price",
            "Total Price",
            "Payment Status",
            "Paid Amount",
            "Date"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[4].height = 20

        # Data rows
        row = 5
        total_price_sum = 0
        paid_amount_sum = 0

        for sale in sales:
            customer_name = customers_map.get(sale.customer_id, sale.customer_id)
            
            # Get line items for this sale
            line_items = db.query(SaleLineItem).filter(SaleLineItem.bill_number == sale.bill_number).all()
            
            if line_items:
                for line_item in line_items:
                    item_name = items_map.get(line_item.item_id, line_item.item_id)
                    
                    cells_data = [
                        sale.bill_number,
                        customer_name,
                        item_name,
                        line_item.quantity,
                        float(line_item.unit_price) if line_item.unit_price else 0,
                        float(line_item.total_price) if line_item.total_price else 0,
                        sale.payment_status,
                        float(sale.paid_amount) if sale.paid_amount else 0,
                        sale.date.strftime("%Y-%m-%d") if sale.date else ""
                    ]

                    for col, value in enumerate(cells_data, 1):
                        cell = ws.cell(row=row, column=col)
                        cell.value = value
                        cell.border = border
                        cell.alignment = Alignment(horizontal='left', vertical='center')

                        # Format currency columns
                        if col in [5, 6, 8]:  # Unit Price, Total Price, Paid Amount
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal='right', vertical='center')

                    # Sum totals
                    total_price_sum += float(line_item.total_price) if line_item.total_price else 0
                    paid_amount_sum += float(sale.paid_amount) if sale.paid_amount else 0
                    row += 1

        # Summary row
        row += 1
        ws[f'A{row}'] = "SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'A{row}'].fill = total_fill
        ws.merge_cells(f'A{row}:I{row}')
        ws.row_dimensions[row].height = 20

        row += 1
        ws[f'A{row}'] = "Total Records:"
        ws[f'B{row}'] = len(sales)
        ws[f'A{row}'].fill = total_fill
        ws[f'B{row}'].fill = total_fill

        row += 1
        ws[f'A{row}'] = "Total Sales Amount:"
        ws[f'B{row}'] = total_price_sum
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'A{row}'].fill = total_fill
        ws[f'B{row}'].fill = total_fill

        row += 1
        ws[f'A{row}'] = "Total Paid:"
        ws[f'B{row}'] = paid_amount_sum
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'A{row}'].fill = total_fill
        ws[f'B{row}'].fill = total_fill

        row += 1
        ws[f'A{row}'] = "Outstanding Balance:"
        ws[f'B{row}'] = total_price_sum - paid_amount_sum
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'A{row}'].fill = total_fill
        ws[f'B{row}'].fill = total_fill

        # Column widths
        column_widths = [28, 20, 20, 12, 12, 15, 15, 12, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Generate filename
        if bill_numbers:
            filename = f"sales-selected-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        else:
            filename = f"sales-all-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export-blow-excel")
async def export_blow_excel(
    blow_ids: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """
    Export blow processes as Excel file.
    If blow_ids provided (comma-separated), export only those.
    Otherwise export all blow processes.
    """
    try:
        # Get blow records
        if blow_ids:
            id_list = [b.strip() for b in blow_ids.split(',')]
            blows = db.query(Blow).filter(Blow.id.in_(id_list)).all()
        else:
            blows = db.query(Blow).all()

        # Get items for lookups
        items_map = {i.id: i.name for i in db.query(Item).all()}

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Blow Processes"

        # Define styles
        header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        title_font = Font(bold=True, size=14, color="FF6B35")
        total_fill = PatternFill(start_color="FFE9DD", end_color="FFE9DD", fill_type="solid")

        # Title
        ws['A1'] = "BLOW PROCESS EXPORT"
        ws['A1'].font = title_font
        ws.merge_cells('A1:I1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25

        # Export date
        export_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A2'] = f"Exported on: {export_date}"
        ws['A2'].font = Font(size=9, italic=True)
        ws.merge_cells('A2:I2')
        ws.row_dimensions[2].height = 18

        # Headers (Row 4)
        headers = [
            "Process ID",
            "From Item (Preform)",
            "To Item (Bottle)",
            "Input Quantity",
            "Output Quantity",
            "Waste Quantity",
            "Efficiency Rate",
            "Cost Per Unit",
            "Date"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[4].height = 20

        # Data rows
        row = 5
        total_cost_sum = 0

        for blow in blows:
            from_item_name = items_map.get(blow.from_item_id, blow.from_item_id)
            to_item_name = items_map.get(blow.to_item_id, blow.to_item_id)
            
            cells_data = [
                blow.id,
                from_item_name,
                to_item_name,
                blow.input_quantity,
                blow.output_quantity,
                blow.waste_quantity,
                float(blow.efficiency_rate) if blow.efficiency_rate else 0,
                float(blow.blow_cost_per_unit) if blow.blow_cost_per_unit else 0,
                blow.date_time.strftime("%Y-%m-%d %H:%M") if blow.date_time else ""
            ]

            for col, value in enumerate(cells_data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')

                # Format currency and percentage columns
                if col == 7:  # Efficiency Rate (%)
                    cell.number_format = '0.00"%"'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col == 8:  # Cost Per Unit
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col in [4, 5, 6]:  # Quantities
                    cell.alignment = Alignment(horizontal='right', vertical='center')

            # Sum total cost
            total_cost_sum += float(blow.blow_cost_per_unit) if blow.blow_cost_per_unit else 0

            row += 1

        # Summary row
        row += 1
        ws[f'A{row}'] = "SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'A{row}'].fill = total_fill
        ws.merge_cells(f'A{row}:I{row}')
        ws.row_dimensions[row].height = 20

        row += 1
        ws[f'A{row}'] = "Total Blow Processes:"
        ws[f'B{row}'] = len(blows)
        ws[f'A{row}'].fill = total_fill
        ws[f'B{row}'].fill = total_fill

        # Column widths
        column_widths = [25, 20, 20, 15, 15, 15, 15, 15, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Generate filename
        if blow_ids:
            filename = f"blow-processes-selected-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        else:
            filename = f"blow-processes-all-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export-waste-excel")
async def export_waste_excel(
    waste_ids: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """
    Export waste records as Excel file.
    If waste_ids provided (comma-separated), export only those.
    Otherwise export all waste records.
    """
    try:
        # Get waste records
        if waste_ids:
            id_list = [w.strip() for w in waste_ids.split(',')]
            wastes = db.query(Waste).filter(Waste.id.in_(id_list)).all()
        else:
            wastes = db.query(Waste).all()

        # Get items for lookups
        items_map = {i.id: i.name for i in db.query(Item).all()}

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Waste Records"

        # Define styles
        header_fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        title_font = Font(bold=True, size=14, color="DC143C")
        total_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

        # Title
        ws['A1'] = "WASTE RECORDS EXPORT"
        ws['A1'].font = title_font
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25

        # Export date
        export_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A2'] = f"Exported on: {export_date}"
        ws['A2'].font = Font(size=9, italic=True)
        ws.merge_cells('A2:H2')
        ws.row_dimensions[2].height = 18

        # Headers (Row 4)
        headers = [
            "Waste ID",
            "Item",
            "Quantity",
            "Recovery Price Per Unit",
            "Total Recovery",
            "Notes",
            "Date",
            "Time"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[4].height = 20

        # Data rows
        row = 5
        total_recovery_sum = 0

        for waste in wastes:
            item_name = items_map.get(waste.item_id, waste.item_id)
            
            cells_data = [
                waste.id,
                item_name,
                waste.quantity,
                float(waste.price_per_unit) if waste.price_per_unit else 0,
                float(waste.total_price) if waste.total_price else 0,
                waste.notes or "",
                waste.date.strftime("%Y-%m-%d") if waste.date else "",
                waste.date.strftime("%H:%M") if waste.date else ""
            ]

            for col, value in enumerate(cells_data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')

                # Format currency and quantity columns
                if col == 3:  # Quantity
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col in [4, 5]:  # Currency columns
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')

            # Sum total recovery
            total_recovery_sum += float(waste.total_price) if waste.total_price else 0

            row += 1

        # Summary row
        row += 1
        ws[f'A{row}'] = "SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'A{row}'].fill = total_fill
        ws.merge_cells(f'A{row}:H{row}')
        ws.row_dimensions[row].height = 20

        row += 1
        ws[f'A{row}'] = "Total Waste Records:"
        ws[f'B{row}'] = len(wastes)
        ws[f'A{row}'].fill = total_fill
        ws[f'B{row}'].fill = total_fill

        row += 1
        ws[f'A{row}'] = "Total Recovery Amount:"
        ws[f'B{row}'] = total_recovery_sum
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'A{row}'].fill = total_fill
        ws[f'B{row}'].fill = total_fill

        # Column widths
        column_widths = [25, 20, 12, 20, 18, 25, 12, 10]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Generate filename
        if waste_ids:
            filename = f"waste-records-selected-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        else:
            filename = f"waste-records-all-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/count/{record_type}")
def get_record_count(record_type: str, db: Session = Depends(get_db)):
    """
    Get the count of records for a specific type to generate incremental IDs.
    Supported types: sale, purchase/purch, blow, waste, exp
    """
    try:
        record_type = record_type.lower()
        if record_type == 'sale':
            count = db.query(Sale).count()
        elif record_type in ['purchase', 'purch']:
            count = db.query(Purchase).count()
        elif record_type == 'blow':
            count = db.query(Blow).count()
        elif record_type == 'waste':
            count = db.query(Waste).count()
        elif record_type == 'exp':
            count = db.query(ExtraExpenditure).count()
        else:
            raise HTTPException(status_code=400, detail=f"Unknown record type: {record_type}")
        
        return {"count": count}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))