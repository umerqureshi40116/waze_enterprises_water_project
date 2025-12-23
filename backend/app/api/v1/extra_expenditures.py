from fastapi import APIRouter, Depends, HTTPException, Body
from sqlalchemy.orm import Session
from typing import List
from datetime import datetime
from app.db.database import get_db
from app.core.security import get_current_user, get_current_admin_user
from app.models.user import User
from app.models.transaction import ExtraExpenditure
from app.schemas.extra_expenditure import ExtraExpenditureCreate, ExtraExpenditureUpdate, ExtraExpenditureResponse
from fastapi.responses import StreamingResponse
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.utils.invoice_pdf_generator import generate_expenditure_invoice_pdf
import asyncio
import logging

logger = logging.getLogger(__name__)

router = APIRouter()


@router.get("/", response_model=List[ExtraExpenditureResponse])
async def get_expenditures(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get all extra expenditures"""
    expenditures = db.query(ExtraExpenditure).order_by(ExtraExpenditure.id.desc()).all()
    return expenditures


@router.post("/", response_model=ExtraExpenditureResponse, status_code=201)
async def create_expenditure(
    expenditure: ExtraExpenditureCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create a new extra expenditure"""
    # Check if ID already exists
    existing = db.query(ExtraExpenditure).filter(ExtraExpenditure.id == expenditure.id).first()
    if existing:
        raise HTTPException(status_code=400, detail="Expenditure ID already exists")

    db_expenditure = ExtraExpenditure(
        **expenditure.dict(),
        created_by=current_user.id
    )
    db.add(db_expenditure)
    db.commit()
    db.refresh(db_expenditure)
    return db_expenditure


# Specific routes must come BEFORE parameterized routes
@router.get("/export/excel")
async def export_expenditures_excel(
    expense_ids: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """Export extra expenditures as Excel file"""
    try:
        # Get expenditures
        if expense_ids:
            id_list = [e.strip() for e in expense_ids.split(',')]
            expenditures = db.query(ExtraExpenditure).filter(ExtraExpenditure.id.in_(id_list)).all()
        else:
            expenditures = db.query(ExtraExpenditure).all()

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Extra Expenditures"

        # Define styles
        header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        title_font = Font(bold=True, size=14, color="8B5CF6")
        total_fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")

        # Title
        ws['A1'] = "EXTRA EXPENDITURES EXPORT"
        ws['A1'].font = title_font
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25

        # Export date
        export_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A2'] = f"Exported on: {export_date}"
        ws['A2'].font = Font(size=9, italic=True)
        ws.merge_cells('A2:G2')
        ws.row_dimensions[2].height = 18

        # Headers (Row 4)
        headers = [
            "Expense ID",
            "Type",
            "Description",
            "Amount (PKR)",
            "Date",
            "Notes",
            "Created By"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[4].height = 20

        # Data rows
        row = 5
        total_amount = 0

        for expenditure in expenditures:
            cells_data = [
                expenditure.id,
                expenditure.expense_type,
                expenditure.description or "",
                float(expenditure.amount) if expenditure.amount else 0,
                expenditure.date.strftime("%Y-%m-%d") if expenditure.date else "",
                expenditure.notes or "",
                expenditure.created_by or ""
            ]

            for col, value in enumerate(cells_data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')

                # Format currency column
                if col == 4:  # Amount
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')

            total_amount += float(expenditure.amount) if expenditure.amount else 0
            row += 1

        # Summary row
        row += 1
        ws[f'A{row}'] = "SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'A{row}'].fill = total_fill
        ws.merge_cells(f'A{row}:G{row}')
        ws.row_dimensions[row].height = 20

        row += 1
        ws[f'A{row}'] = "Total Records:"
        ws[f'B{row}'] = len(expenditures)
        ws[f'A{row}'].fill = total_fill
        ws[f'B{row}'].fill = total_fill

        row += 1
        ws[f'A{row}'] = "Total Amount:"
        ws[f'B{row}'] = total_amount
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'A{row}'].fill = total_fill
        ws[f'B{row}'].fill = total_fill

        # Column widths
        column_widths = [18, 18, 25, 15, 12, 20, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Generate filename
        if expense_ids:
            filename = f"extra-expenditures-selected-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        else:
            filename = f"extra-expenditures-all-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/total")
async def get_total_expenditures(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get total extra expenditures for current month"""
    from sqlalchemy import extract, func
    from datetime import datetime as dt

    current_month = dt.now().month
    current_year = dt.now().year

    total = db.query(func.sum(ExtraExpenditure.amount)).filter(
        extract('month', ExtraExpenditure.date) == current_month,
        extract('year', ExtraExpenditure.date) == current_year
    ).scalar()

    count = db.query(func.count(ExtraExpenditure.id)).filter(
        extract('month', ExtraExpenditure.date) == current_month,
        extract('year', ExtraExpenditure.date) == current_year
    ).scalar()

    return {
        "total_amount": float(total) if total else 0.0,
        "count": count or 0
    }


# Parameterized routes come LAST
@router.put("/{expenditure_id}", response_model=ExtraExpenditureResponse)
async def update_expenditure(
    expenditure_id: str,
    expenditure: ExtraExpenditureUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """Update an extra expenditure (admin only)"""
    db_expenditure = db.query(ExtraExpenditure).filter(ExtraExpenditure.id == expenditure_id).first()
    if not db_expenditure:
        raise HTTPException(status_code=404, detail="Expenditure not found")

    update_data = expenditure.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_expenditure, field, value)

    db.commit()
    db.refresh(db_expenditure)
    return db_expenditure


@router.delete("/{expenditure_id}")
async def delete_expenditure(
    expenditure_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """Delete an extra expenditure (admin only)"""
    db_expenditure = db.query(ExtraExpenditure).filter(ExtraExpenditure.id == expenditure_id).first()
    if not db_expenditure:
        raise HTTPException(status_code=404, detail="Expenditure not found")

    db.delete(db_expenditure)
    db.commit()
    return {"message": "Expenditure deleted successfully"}


@router.get("/pdf/{expenditure_id}")
async def download_expenditure_invoice(
    expenditure_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Download operating expense record as PDF invoice"""
    try:
        # Get expenditure record
        expenditure = db.query(ExtraExpenditure).filter(ExtraExpenditure.id == expenditure_id).first()
        if not expenditure:
            raise HTTPException(status_code=404, detail="Expenditure record not found")
        
        # Generate expenditure report PDF in a thread
        pdf_buffer = await asyncio.to_thread(
            generate_expenditure_invoice_pdf,
            expenditure=expenditure
        )
        
        filename = f"expense_record_{expenditure_id}_{expenditure.date.strftime('%Y%m%d') if hasattr(expenditure, 'date') and expenditure.date else 'unknown'}.pdf"
        
        return StreamingResponse(
            iter([pdf_buffer.getvalue()]),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating expenditure invoice PDF: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")
